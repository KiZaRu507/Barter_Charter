"""
excel_logger.py

Excel logging for the Barter Charter simulation using openpyxl.

Creates and maintains an Excel file with three sheets:

1) Commodities
   - Round
   - Commodity
   - PriceRs
   - RatioVsBase
   - MinUnits
   - MaxUnits

2) Trades
   - TradeID
   - Round
   - FromTeam
   - ToTeam
   - GiveCommodity
   - GiveQty
   - ReceiveCommodity
   - ReceiveQty

3) Portfolios
   - Round
   - Team
   - TotalValueRs
   - TotalValueBaseUnits
   - <CommodityName>_units (one column per commodity)
"""

from typing import Dict
from openpyxl import Workbook

from game_engine import Commodity, GameState


class ExcelLogger:
    """
    Excel logger using openpyxl.

    Usage pattern:
        logger = ExcelLogger("barter_charter.xlsx")
        logger.log_commodities(game_state.commodities, round_no=0)
        logger.log_portfolios_round(game_state)
        logger.log_trade(trade)
    """

    def __init__(self, filename: str = "barter_charter.xlsx"):
        self.filename = filename
        self.wb = Workbook()

        # Create sheets
        self.sheet_commodities = self.wb.active
        self.sheet_commodities.title = "Commodities"
        self.sheet_trades = self.wb.create_sheet("Trades")
        self.sheet_portfolios = self.wb.create_sheet("Portfolios")

        # Headers for Commodities sheet
        self.sheet_commodities.append(
            ["Round", "Commodity", "PriceRs", "RatioVsBase", "MinUnits", "MaxUnits"]
        )

        # Headers for Trades sheet
        self.sheet_trades.append([
            "TradeID", "Round", "FromTeam", "ToTeam",
            "GiveCommodity", "GiveQty", "ReceiveCommodity", "ReceiveQty"
        ])

        # Headers for Portfolios sheet
        self.sheet_portfolios.append(["Round", "Team", "TotalValueRs", "TotalValueBaseUnits"])

        # Internal counter for TradeID
        self.trade_counter = 0

        # Save initial empty structure
        self.save()

    # -----------------------------------------------------
    # Core helper
    # -----------------------------------------------------

    def save(self):
        """
        Save the workbook to disk.
        """
        self.wb.save(self.filename)

    # -----------------------------------------------------
    # Commodities logging
    # -----------------------------------------------------

    def log_commodities(self, commodities: Dict[str, Commodity], round_no: int):
        """
        Append one row per commodity for the given round.
        """
        for c in commodities.values():
            self.sheet_commodities.append(
                [round_no, c.name, c.price, c.base_ratio, c.min_units, c.max_units]
            )
        self.save()

    # -----------------------------------------------------
    # Portfolio logging
    # -----------------------------------------------------

    def ensure_portfolio_commodity_columns(self, commodities: Dict[str, Commodity]):
        """
        Ensure the 'Portfolios' sheet has a _units column for each commodity.
        If new commodities are added later, columns will be appended.

        Columns will be named: <CommodityName>_units
        """
        header_row = self.sheet_portfolios[1]
        headers = [cell.value for cell in header_row]

        for cname in commodities.keys():
            col_name = f"{cname}_units"
            if col_name not in headers:
                col_index = self.sheet_portfolios.max_column + 1
                self.sheet_portfolios.cell(row=1, column=col_index, value=col_name)
                headers.append(col_name)

        self.save()

    def log_portfolios_round(self, game_state: GameState):
        """
        Append one row per team for the current round.

        Columns:
        - Round
        - Team
        - TotalValueRs
        - TotalValueBaseUnits
        - <Commodity>_units...
        """
        round_no = game_state.current_round
        commodities = game_state.commodities
        base = game_state.base_commodity

        # Ensure columns exist for all commodities
        self.ensure_portfolio_commodity_columns(commodities)

        # Build mapping from commodity name to column index
        header_row = self.sheet_portfolios[1]
        headers = [cell.value for cell in header_row]
        commodity_cols = {}
        for cname in commodities.keys():
            col_name = f"{cname}_units"
            if col_name in headers:
                commodity_cols[cname] = headers.index(col_name) + 1

        # Append a row for each team
        for team in game_state.teams.values():
            row_idx = self.sheet_portfolios.max_row + 1
            total_rs = team.value_rs(commodities)
            total_base = team.value_in_base(commodities, base)

            self.sheet_portfolios.cell(row=row_idx, column=1, value=round_no)
            self.sheet_portfolios.cell(row=row_idx, column=2, value=team.name)
            self.sheet_portfolios.cell(row=row_idx, column=3, value=total_rs)
            self.sheet_portfolios.cell(row=row_idx, column=4, value=total_base)

            # Commodity units
            for cname, col_idx in commodity_cols.items():
                units = team.holdings.get(cname, 0)
                self.sheet_portfolios.cell(row=row_idx, column=col_idx, value=units)

        self.save()

    # -----------------------------------------------------
    # Trades logging
    # -----------------------------------------------------

    def log_trade(self, trade):
        """
        Log a trade in the Trades sheet.

        Simplified assumption: we only log the first entry from
        'give' and 'receive' dicts, since in the current design
        there's exactly one give-commodity and one receive-commodity.
        """
        self.trade_counter += 1

        give_name, give_qty = next(iter(trade.give.items())) if trade.give else ("", 0)
        recv_name, recv_qty = next(iter(trade.receive.items())) if trade.receive else ("", 0)

        self.sheet_trades.append([
            self.trade_counter,
            trade.round_no,
            trade.from_team,
            trade.to_team,
            give_name,
            give_qty,
            recv_name,
            recv_qty
        ])

        self.save()
